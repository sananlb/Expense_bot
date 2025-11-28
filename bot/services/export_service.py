"""
Сервис для экспорта финансовых данных в CSV и XLSX форматы.
"""
import csv
import calendar
from io import StringIO, BytesIO
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from decimal import Decimal
from typing import List, Dict, Any, Union, Tuple
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.label import DataLabel, DataLabelList
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.formatting.rule import FormulaRule
try:
    from openpyxl.chart.text import RichText, Paragraph
    from openpyxl.drawing.text import ParagraphProperties, CharacterProperties
except ImportError:
    RichText = Paragraph = ParagraphProperties = CharacterProperties = None

from expenses.models import Expense, Income, Cashback, Profile
from bot.utils.language import get_text

logger = logging.getLogger(__name__)


CM_TO_EMU = 360000  # 1 сантиметр в единицах EMU (Excel)


def smart_number(value) -> Union[int, float]:
    """
    Конвертирует число для Excel: целые как int, дробные как float.
    Это гарантирует корректное отображение без лишних десятичных знаков.
    """
    if value is None:
        return 0
    if isinstance(value, Decimal):
        value = float(value)
    if isinstance(value, (int, float)):
        if value == int(value):
            return int(value)
        return round(value, 2)
    return value


class ExportService:
    """Сервис для экспорта данных о тратах и доходах"""

    # Цветовая палитра для категорий (приятные приглушенные цвета)
    CATEGORY_COLORS = [
        '#5B9BD5',  # мягкий синий
        '#ED7D31',  # теплый оранжевый
        '#70AD47',  # приятный зеленый
        '#9966CC',  # мягкий фиолетовый
        '#FFC000',  # золотистый
        '#44B5E2',  # небесный голубой
        '#C55A9A',  # пыльная роза
        '#7FB685',  # мятный зеленый
        '#E07B91',  # коралловый
        '#CC8D66',  # терракотовый
        '#6FA8DC',  # светло-синий
        '#D9A441',  # охра
        '#B19CD9',  # лавандовый
        '#5FA3A3',  # бирюзовый
        '#D97BA6'   # фуксия приглушенная
    ]

    @staticmethod
    def calculate_category_cashbacks(expenses: List[Expense], user_id: int, month: int, household_mode: bool = False) -> Dict:
        """
        Рассчитать кешбэк по категориям.

        Args:
            expenses: Список трат
            user_id: ID пользователя
            month: Месяц для которого считаем кешбэк
            household_mode: Если True, учитываем кешбэк всех членов семьи

        Returns:
            Словарь {category_id: cashback_amount}
        """
        try:
            # Получаем профиль пользователя
            profile = Profile.objects.get(telegram_id=user_id)

            # Получаем весь кешбэк (пользователя или всей семьи)
            if household_mode and profile.household:
                # В household mode берем кешбэк всех членов семьи
                user_cashbacks = list(Cashback.objects.filter(
                    profile__household=profile.household,
                    month=month
                ).select_related('category', 'profile'))
            else:
                # В личном режиме только кешбэк текущего пользователя
                user_cashbacks = list(Cashback.objects.filter(
                    profile=profile,
                    month=month
                ).select_related('category'))

            # Создаем словарь кешбеков по категориям
            cashback_by_category = {}
            for cb in user_cashbacks:
                if cb.category_id:
                    if cb.category_id not in cashback_by_category:
                        cashback_by_category[cb.category_id] = []
                    cashback_by_category[cb.category_id].append(cb)

            # Подсчитываем траты по категориям
            category_amounts = {}
            for expense in expenses:
                if expense.category_id:
                    if expense.category_id not in category_amounts:
                        category_amounts[expense.category_id] = 0
                    category_amounts[expense.category_id] += float(expense.amount)

            # Рассчитываем кешбек для каждой категории
            category_cashbacks = {}
            for category_id, amount in category_amounts.items():
                cashback = 0
                if category_id in cashback_by_category:
                    for cb in cashback_by_category[category_id]:
                        cb_amount = amount
                        if cb.limit_amount and cb.limit_amount > 0:
                            cb_amount = min(amount, float(cb.limit_amount))
                        cashback += cb_amount * (float(cb.cashback_percent) / 100)
                category_cashbacks[category_id] = cashback

            return category_cashbacks
        except Exception as e:
            logger.error(f"Error calculating cashbacks: {e}")
            return {}

    @staticmethod
    def prepare_operations_data(
        expenses: List[Expense],
        incomes: List[Income],
        lang: str = 'ru'
    ) -> List[Dict[str, Any]]:
        """
        Подготовить данные операций для экспорта.

        Args:
            expenses: Список трат
            incomes: Список доходов
            lang: Язык для категорий (ru/en)

        Returns:
            Список словарей с данными операций, отсортированный по дате (от новых к старым)
        """
        operations = []

        # Добавить траты
        for expense in expenses:
            operations.append({
                'date': expense.expense_date,
                'time': expense.expense_time or expense.created_at.time(),  # Fallback на время создания
                'type': 'expense',
                'amount': -float(expense.amount),  # Отрицательное для трат
                'currency': expense.currency,
                'category': expense.category.get_display_name(lang) if expense.category else get_text('no_category', lang),
                'category_id': expense.category_id,  # ID категории для расчета кешбэка
                'description': expense.description or '',
                'object': expense  # Сохраняем объект для доступа к дополнительным полям
            })

        # Добавить доходы
        for income in incomes:
            operations.append({
                'date': income.income_date,
                'time': income.income_time or income.created_at.time(),  # Fallback на время создания
                'type': 'income',
                'amount': float(income.amount),  # Положительное для доходов
                'currency': income.currency,
                'category': income.category.get_display_name(lang) if income.category else get_text('no_category', lang),
                'category_id': income.category_id,  # ID категории
                'description': income.description or '',
                'object': income  # Сохраняем объект для доступа к дополнительным полям
            })

        # Сортировать от новых к старым
        operations.sort(key=lambda x: (x['date'], x['time']), reverse=True)

        return operations

    @staticmethod
    def generate_csv(
        expenses: List[Expense],
        incomes: List[Income],
        year: int,
        month: int,
        lang: str = 'ru',
        user_id: int = None,
        household_mode: bool = False
    ) -> bytes:
        """
        Генерация CSV файла с операциями за месяц.

        Args:
            expenses: Список трат
            incomes: Список доходов
            year: Год
            month: Месяц
            lang: Язык (ru/en)
            user_id: ID пользователя для расчета кешбэка
            household_mode: Режим семейного бюджета

        Returns:
            Байты CSV файла (UTF-8 с BOM для корректного открытия в Excel)
        """
        operations = ExportService.prepare_operations_data(expenses, incomes, lang)

        # Рассчитываем кешбэк для каждой траты
        expense_cashbacks = {}  # {expense_id: cashback_amount}
        has_any_cashback = False

        if user_id and expenses:
            category_cashbacks = ExportService.calculate_category_cashbacks(expenses, user_id, month, household_mode)

            # Рассчитываем кешбэк для каждой траты
            for expense in expenses:
                if expense.category_id and expense.category_id in category_cashbacks:
                    cashback_amount = float(expense.amount) * (category_cashbacks[expense.category_id] / sum(
                        float(e.amount) for e in expenses if e.category_id == expense.category_id
                    ))
                    expense_cashbacks[expense.id] = cashback_amount
                    if cashback_amount > 0:
                        has_any_cashback = True

        output = StringIO()
        # Добавляем BOM в начало для корректного открытия в Excel
        output.write('\ufeff')

        # Заголовки в зависимости от языка
        # Порядок: Дата, Время, Сумма, Кешбэк (если есть), Валюта, Категория, Описание, Тип
        if has_any_cashback:
            if lang == 'en':
                headers = ['Date', 'Time', 'Amount', 'Cashback', 'Currency', 'Category', 'Description', 'Type']
            else:
                headers = ['Дата', 'Время', 'Сумма', 'Кешбэк', 'Валюта', 'Категория', 'Описание', 'Тип']
        else:
            if lang == 'en':
                headers = ['Date', 'Time', 'Amount', 'Currency', 'Category', 'Description', 'Type']
            else:
                headers = ['Дата', 'Время', 'Сумма', 'Валюта', 'Категория', 'Описание', 'Тип']

        writer = csv.writer(output, delimiter=',', quotechar='"', quoting=csv.QUOTE_ALL)
        writer.writerow(headers)

        # Данные
        for op in operations:
            type_text = get_text('income', lang) if op['type'] == 'income' else get_text('expense', lang)

            # Санитизируем описание и категорию - убираем переносы строк и лишние пробелы
            description = str(op['description'] or '')
            description = description.replace('\n', ' ').replace('\r', ' ').strip()

            category = str(op['category'] or '')
            category = category.replace('\n', ' ').replace('\r', ' ').strip()

            # Получаем кешбэк для траты
            cashback = 0
            if op['type'] == 'expense' and 'object' in op and hasattr(op['object'], 'id'):
                cashback = expense_cashbacks.get(op['object'].id, 0)

            # Формируем строку в зависимости от наличия колонки кешбэка
            if has_any_cashback:
                writer.writerow([
                    op['date'].strftime('%d.%m.%Y'),
                    op['time'].strftime('%H:%M'),
                    f"{op['amount']:.2f}",
                    f"{cashback:.2f}" if op['type'] == 'expense' else '',
                    op['currency'],
                    category,
                    description,
                    type_text
                ])
            else:
                writer.writerow([
                    op['date'].strftime('%d.%m.%Y'),
                    op['time'].strftime('%H:%M'),
                    f"{op['amount']:.2f}",
                    op['currency'],
                    category,
                    description,
                    type_text
                ])

        # Вернуть байты (BOM уже добавлен в начало StringIO)
        return output.getvalue().encode('utf-8')

    @staticmethod
    def _get_month_sheet_name(year: int, month: int, lang: str) -> str:
        """Возвращает название листа в формате 'Ноя-2024' или 'Nov-2024'."""
        month_names_en = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                         'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        month_names_ru = ['Янв', 'Фев', 'Мар', 'Апр', 'Май', 'Июн',
                         'Июл', 'Авг', 'Сен', 'Окт', 'Ноя', 'Дек']
        if lang == 'ru':
            return f"{month_names_ru[month-1]}-{year}"
        else:
            return f"{month_names_en[month-1]}-{year}"

    @staticmethod
    def _load_12_months_data(
        profile_id: int,
        current_year: int,
        current_month: int,
        lang: str,
        household_id: int | None = None
    ) -> Tuple[Dict[Tuple[int, int], Dict[str, float]], Dict[Tuple[int, int], Dict[str, float]], Dict[Tuple[int, int], Dict[str, int]], List[Tuple[int, int]], Dict[Tuple[str, str], Dict[str, Any]], Dict[Tuple[str, str], Dict[str, Any]]]:
        """
        Загружает данные о расходах и доходах за последние 12 месяцев.

        Returns:
            Tuple из:
            - expenses_by_month: {(year, month): {category_name: total_amount}}
            - incomes_by_month: {(year, month): {category_name: total_amount}}
            - expenses_counts_by_month: {(year, month): {category_name: operations_count}}
            - months_list: список (year, month) за 12 месяцев от старого к новому
        """
        # Формируем список 12 месяцев (от старого к новому)
        months_list = []
        current_date = date(current_year, current_month, 1)
        for i in range(11, -1, -1):  # 11 месяцев назад до текущего
            month_date = current_date - relativedelta(months=i)
            months_list.append((month_date.year, month_date.month))
        month_index = {m: idx for idx, m in enumerate(months_list)}
        month_index = {m: idx for idx, m in enumerate(months_list)}

        # Определяем диапазон дат
        start_date = date(months_list[0][0], months_list[0][1], 1)
        end_year, end_month = months_list[-1]
        last_day = calendar.monthrange(end_year, end_month)[1]
        end_date = date(end_year, end_month, last_day)

        # Загружаем расходы
        if household_id:
            expenses = Expense.objects.filter(
                profile__household_id=household_id,
                expense_date__gte=start_date,
                expense_date__lte=end_date
            ).select_related('category')
        else:
            expenses = Expense.objects.filter(
                profile_id=profile_id,
                expense_date__gte=start_date,
                expense_date__lte=end_date
            ).select_related('category')

        # Группируем расходы по месяцам и категориям + по операциям (описание+валюта)
        expenses_by_month: Dict[Tuple[int, int], Dict[str, float]] = {m: {} for m in months_list}
        expenses_counts_by_month: Dict[Tuple[int, int], Dict[str, int]] = {m: {} for m in months_list}
        expense_operations: Dict[Tuple[str, str], Dict[str, Any]] = {}
        for expense in expenses:
            year_month = (expense.expense_date.year, expense.expense_date.month)
            category_name = expense.category.get_display_name(lang) if expense.category else 'Без категории'
            if year_month in expenses_by_month:
                if category_name not in expenses_by_month[year_month]:
                    expenses_by_month[year_month][category_name] = 0
                    expenses_counts_by_month[year_month][category_name] = 0
                expenses_by_month[year_month][category_name] += float(expense.amount)
                expenses_counts_by_month[year_month][category_name] += 1

            desc_display = (expense.description or 'Без описания').strip()
            desc_norm = desc_display.lower()
            key = (desc_norm, expense.currency)
            if key not in expense_operations:
                expense_operations[key] = {
                    'description': desc_display if desc_display else 'Без описания',
                    'category': category_name,
                    'currency': expense.currency,
                    'monthly_totals': [0 for _ in months_list],
                    'total': 0.0,
                    'count': 0,
                }
            op = expense_operations[key]
            idx = month_index.get(year_month)
            if idx is not None:
                op['monthly_totals'][idx] += float(expense.amount)
            op['total'] += float(expense.amount)
            op['count'] += 1

        # Загружаем доходы
        if household_id:
            incomes = Income.objects.filter(
                profile__household_id=household_id,
                income_date__gte=start_date,
                income_date__lte=end_date
            ).select_related('category')
        else:
            incomes = Income.objects.filter(
                profile_id=profile_id,
                income_date__gte=start_date,
                income_date__lte=end_date
            ).select_related('category')

        # Группируем доходы по месяцам и категориям + по операциям (описание+валюта)
        incomes_by_month: Dict[Tuple[int, int], Dict[str, float]] = {m: {} for m in months_list}
        income_operations: Dict[Tuple[str, str], Dict[str, Any]] = {}
        for income in incomes:
            year_month = (income.income_date.year, income.income_date.month)
            category_name = income.category.get_display_name(lang) if income.category else 'Без категории'
            if year_month in incomes_by_month:
                if category_name not in incomes_by_month[year_month]:
                    incomes_by_month[year_month][category_name] = 0
                incomes_by_month[year_month][category_name] += float(income.amount)

            desc_display = (income.description or 'Без описания').strip()
            desc_norm = desc_display.lower()
            key = (desc_norm, income.currency)
            if key not in income_operations:
                income_operations[key] = {
                    'description': desc_display if desc_display else 'Без описания',
                    'category': category_name,
                    'currency': income.currency,
                    'monthly_totals': [0 for _ in months_list],
                    'total': 0.0,
                    'count': 0,
                }
            op = income_operations[key]
            idx = month_index.get(year_month)
            if idx is not None:
                op['monthly_totals'][idx] += float(income.amount)
            op['total'] += float(income.amount)
            op['count'] += 1

        return expenses_by_month, incomes_by_month, expenses_counts_by_month, months_list, expense_operations, income_operations

    @staticmethod
    def _add_summary_sheet(
        wb: Workbook,
        profile_id: int,
        current_year: int,
        current_month: int,
        lang: str,
        household_id: int | None = None
    ) -> Dict[str, Any]:
        """
        Добавляет лист "Итоги 12 мес" в workbook с данными за последние 12 месяцев.
        Текущий месяц содержит ссылки на лист с детальным отчётом.
        """
        # Загружаем данные
        expenses_by_month, incomes_by_month, expenses_counts_by_month, months_list, expense_operations, income_operations = ExportService._load_12_months_data(
            profile_id, current_year, current_month, lang, household_id
        )

        # Функция для создания Excel формулы тренда (линейная регрессия)
        def make_trend_formula(first_col: str, last_col: str, row: int, num_months: int) -> str:
            """
            Создаёт Excel формулу для расчёта годового тренда через SLOPE.
            Формула: SLOPE(данные, {1,2,...,n}) * 12 / AVERAGE(данные)
            Результат: % изменения за год на основе линейного тренда
            """
            # Создаём массив индексов {1,2,3,...,n}
            indices = ",".join(str(i) for i in range(1, num_months + 1))
            data_range = f"{first_col}{row}:{last_col}{row}"
            # IFERROR на случай если все значения 0 или нет данных
            return f'=IFERROR(SLOPE({data_range},{{{indices}}})*12/AVERAGE({data_range}),"")'

        def calc_trend_python(values: list[float]) -> float | None:
            """
            Python версия расчёта тренда (линейная регрессия).
            Используется для аналитических таблиц где данные уже в Python.
            Возвращает: % изменения за год на основе линейного тренда.
            """
            # Фильтруем None, но оставляем 0 (важно для тренда)
            clean_values = [v if v is not None else 0 for v in values]
            n = len(clean_values)
            if n < 2:
                return None

            # Проверяем есть ли хоть какие-то данные
            if all(v == 0 for v in clean_values):
                return None

            # Линейная регрессия: slope = Σ(x-x̄)(y-ȳ) / Σ(x-x̄)²
            x_mean = (n - 1) / 2  # среднее индексов 0,1,2,...,n-1
            y_mean = sum(clean_values) / n

            numerator = sum((i - x_mean) * (v - y_mean) for i, v in enumerate(clean_values))
            denominator = sum((i - x_mean) ** 2 for i in range(n))

            if denominator == 0 or y_mean == 0:
                return None

            slope = numerator / denominator
            # Экстраполяция на год и нормализация к среднему
            annual_change = (slope * 12) / y_mean
            return annual_change

        expense_month_totals = [sum(expenses_by_month[(y, m)].values()) for y, m in months_list]
        income_month_totals = [sum(incomes_by_month[(y, m)].values()) for y, m in months_list]

        # Создаём лист (второй после листа месяца)
        sheet_title = "Итоги 12 мес" if lang == 'ru' else "Summary 12m"
        ws = wb.create_sheet(title=sheet_title)  # Добавляется в конец

        # Стили
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        income_header_fill = PatternFill(start_color="7FB685", end_color="7FB685", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Названия месяцев
        month_names_en = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                         'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        month_names_ru = ['Янв', 'Фев', 'Мар', 'Апр', 'Май', 'Июн',
                         'Июл', 'Авг', 'Сен', 'Окт', 'Ноя', 'Дек']

        def get_month_header(year: int, month: int) -> str:
            """Возвращает заголовок месяца, например 'Ноя'24'."""
            if lang == 'ru':
                return f"{month_names_ru[month-1]}'{str(year)[2:]}"
            else:
                return f"{month_names_en[month-1]}'{str(year)[2:]}"

        # ==================== СЕКЦИЯ РАСХОДОВ ====================
        # Заголовок секции
        section_title = 'РАСХОДЫ ЗА 12 МЕСЯЦЕВ' if lang == 'ru' else 'EXPENSES FOR 12 MONTHS'
        title_cell = ws.cell(row=1, column=1, value=section_title)
        title_cell.font = Font(bold=True, color="FFFFFF", size=12)
        title_cell.fill = PatternFill(start_color="C85A54", end_color="C85A54", fill_type="solid")
        title_cell.alignment = header_alignment
        title_cell.border = thin_border

        # Заголовки колонок: Категория | План | Мес1 | Мес2 | ... | Мес12 | Итого | Среднее | Изм. %
        headers = ['Категория' if lang == 'ru' else 'Category',
                   'План' if lang == 'ru' else 'Plan']
        for year, month in months_list:
            headers.append(get_month_header(year, month))
        headers.append('Итого' if lang == 'ru' else 'Total')
        headers.append('Среднее' if lang == 'ru' else 'Average')
        headers.append('Изм. %' if lang == 'ru' else 'Change %')

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Собираем все категории расходов за 12 месяцев
        all_expense_categories = set()
        for month_data in expenses_by_month.values():
            all_expense_categories.update(month_data.keys())

        # Сортируем по общей сумме (от большего к меньшему)
        category_totals = {}
        for cat in all_expense_categories:
            total = sum(expenses_by_month[m].get(cat, 0) for m in months_list)
            category_totals[cat] = total
        sorted_categories = sorted(all_expense_categories, key=lambda c: category_totals[c], reverse=True)

        # Заполняем данные по категориям расходов
        current_row = 3
        current_sheet_name = ExportService._get_month_sheet_name(current_year, current_month, lang)

        # Словарь для хранения строк категорий (для ссылок)
        expense_category_rows = {}  # {category_name: row_number}

        # Колонки для формул
        total_col = 3 + len(months_list)  # Итого
        avg_col = total_col + 1  # Среднее
        change_col = avg_col + 1  # Изменение %
        first_month_col = 3  # Первая колонка с месяцем (C)
        last_month_col = 2 + len(months_list)  # Последняя колонка с месяцем

        # Вычисляем номер строки ИТОГО РАСХОДЫ заранее (для формул среднего)
        # Строка 3 = первая категория, +len(categories) = строка после последней категории = ИТОГО
        expenses_totals_row_num = 3 + len(sorted_categories)

        for category in sorted_categories:
            row_values: list[float] = []
            ws.cell(row=current_row, column=1, value=category)
            ws.cell(row=current_row, column=2, value=None)  # План - пустой

            for col_offset, (year, month) in enumerate(months_list):
                amount = expenses_by_month[(year, month)].get(category, 0)
                col_idx = 3 + col_offset
                is_current_month = (year == current_year and month == current_month)

                if is_current_month:
                    # Текущий месяц = формула SUMIF со ссылкой на лист месяца
                    # K = категория, M = всего (summary_start_col=11, +2=13)
                    formula = f"=SUMIF('{current_sheet_name}'!K:K,A{current_row},'{current_sheet_name}'!M:M)"
                    cell = ws.cell(row=current_row, column=col_idx, value=formula)
                    cell.font = Font(color="0563C1")  # Синий цвет для формул-ссылок
                elif amount > 0:
                    ws.cell(row=current_row, column=col_idx, value=smart_number(amount))
                else:
                    ws.cell(row=current_row, column=col_idx, value=None)
                row_values.append(float(amount) if amount else 0)

            # Итого = формула SUM по всем месяцам
            first_col_letter = get_column_letter(first_month_col)
            last_col_letter = get_column_letter(last_month_col)
            sum_formula = f"=SUM({first_col_letter}{current_row}:{last_col_letter}{current_row})"
            ws.cell(row=current_row, column=total_col, value=sum_formula)

            # Среднее = Итого / количество месяцев где ИТОГО РАСХОДЫ > 0
            # Формула: O3 / COUNTIF(C20:N20, "<>0") - где 20 = строка ИТОГО РАСХОДЫ
            total_col_letter = get_column_letter(total_col)
            avg_formula = f"=IFERROR({total_col_letter}{current_row}/COUNTIF({first_col_letter}{expenses_totals_row_num}:{last_col_letter}{expenses_totals_row_num},\"<>0\"),\"\")"
            ws.cell(row=current_row, column=avg_col, value=avg_formula)

            # Изменение = Excel формула тренда (SLOPE)
            trend_formula = make_trend_formula(first_col_letter, last_col_letter, current_row, len(months_list))
            change_cell = ws.cell(row=current_row, column=change_col, value=trend_formula)
            change_cell.number_format = '0.0%'

            # Применяем границы и чередование
            is_even = (current_row % 2 == 0)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border
                if is_even:
                    cell.fill = gray_fill

            expense_category_rows[category] = current_row
            current_row += 1

        # Пустая строка перед ИТОГО
        last_data_row = current_row - 1  # Последняя строка с данными категорий
        current_row += 1  # Пропускаем строку

        # Строка ИТОГО РАСХОДЫ
        expenses_totals_row = current_row
        first_data_row = 3  # Первая строка с данными категорий

        ws.cell(row=current_row, column=1, value='ИТОГО РАСХОДЫ' if lang == 'ru' else 'TOTAL EXPENSES')
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=None)  # План

        # Для каждого месяца - формула SUM по колонке
        for col_offset, (year, month) in enumerate(months_list):
            col_idx = 3 + col_offset
            col_letter = get_column_letter(col_idx)
            is_current_month = (year == current_year and month == current_month)

            # Формула суммы по колонке
            sum_col_formula = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
            cell = ws.cell(row=current_row, column=col_idx, value=sum_col_formula)
            if is_current_month:
                cell.font = Font(bold=True, color="0563C1")
            else:
                cell.font = Font(bold=True)

        # Итого = SUM по строке (все месяцы)
        first_col_letter = get_column_letter(first_month_col)
        last_col_letter = get_column_letter(last_month_col)
        sum_formula = f"=SUM({first_col_letter}{current_row}:{last_col_letter}{current_row})"
        ws.cell(row=current_row, column=total_col, value=sum_formula)
        ws.cell(row=current_row, column=total_col).font = Font(bold=True)

        # Среднее для ИТОГО = Итого / COUNTIF по своей строке (считаем непустые месяцы)
        total_col_letter = get_column_letter(total_col)
        avg_formula = f"=IFERROR({total_col_letter}{current_row}/COUNTIF({first_col_letter}{current_row}:{last_col_letter}{current_row},\"<>0\"),\"\")"
        ws.cell(row=current_row, column=avg_col, value=avg_formula)
        ws.cell(row=current_row, column=avg_col).font = Font(bold=True)

        # Изменение % для ИТОГО РАСХОДЫ (Excel формула тренда)
        trend_formula = make_trend_formula(first_col_letter, last_col_letter, current_row, len(months_list))
        change_cell = ws.cell(row=current_row, column=change_col, value=trend_formula)
        change_cell.font = Font(bold=True)
        change_cell.number_format = '0.0%'

        # Границы для строки итого
        for col in range(1, len(headers) + 1):
            ws.cell(row=current_row, column=col).border = thin_border

        current_row += 2  # Пустая строка перед доходами

        # ==================== СЕКЦИЯ ДОХОДОВ ====================
        # Заголовок секции доходов
        income_title = 'ДОХОДЫ ЗА 12 МЕСЯЦЕВ' if lang == 'ru' else 'INCOME FOR 12 MONTHS'
        title_cell = ws.cell(row=current_row, column=1, value=income_title)
        title_cell.font = Font(bold=True, color="FFFFFF", size=12)
        title_cell.fill = PatternFill(start_color="5FA85F", end_color="5FA85F", fill_type="solid")
        title_cell.alignment = header_alignment
        title_cell.border = thin_border
        current_row += 1

        # Заголовки для доходов
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="7FB685", end_color="7FB685", fill_type="solid")
            cell.alignment = header_alignment
            cell.border = thin_border
        current_row += 1

        # Собираем все категории доходов
        all_income_categories = set()
        for month_data in incomes_by_month.values():
            all_income_categories.update(month_data.keys())

        # Сортируем по общей сумме
        income_totals = {}
        for cat in all_income_categories:
            total = sum(incomes_by_month[m].get(cat, 0) for m in months_list)
            income_totals[cat] = total
        sorted_income_categories = sorted(all_income_categories, key=lambda c: income_totals[c], reverse=True)

        # Заполняем данные по категориям доходов
        income_category_rows = {}
        income_first_data_row = current_row  # Запоминаем первую строку доходов

        # Вычисляем номер строки ИТОГО ДОХОДЫ заранее (для формул среднего)
        income_totals_row_num = income_first_data_row + len(sorted_income_categories)

        for category in sorted_income_categories:
            row_values: list[float] = []
            ws.cell(row=current_row, column=1, value=category)
            ws.cell(row=current_row, column=2, value=None)  # План

            for col_offset, (year, month) in enumerate(months_list):
                amount = incomes_by_month[(year, month)].get(category, 0)
                col_idx = 3 + col_offset
                is_current_month = (year == current_year and month == current_month)

                if is_current_month:
                    # Текущий месяц = формула SUMIF со ссылкой на лист месяца (доходы тоже в K:K, M:M)
                    formula = f"=SUMIF('{current_sheet_name}'!K:K,A{current_row},'{current_sheet_name}'!M:M)"
                    cell = ws.cell(row=current_row, column=col_idx, value=formula)
                    cell.font = Font(color="0563C1")  # Синий цвет для формул-ссылок
                elif amount > 0:
                    ws.cell(row=current_row, column=col_idx, value=smart_number(amount))
                else:
                    ws.cell(row=current_row, column=col_idx, value=None)
                row_values.append(float(amount) if amount else 0)

            # Итого = формула SUM по всем месяцам
            first_col_letter = get_column_letter(first_month_col)
            last_col_letter = get_column_letter(last_month_col)
            sum_formula = f"=SUM({first_col_letter}{current_row}:{last_col_letter}{current_row})"
            ws.cell(row=current_row, column=total_col, value=sum_formula)

            # Среднее = Итого / количество месяцев где ИТОГО ДОХОДЫ > 0
            total_col_letter = get_column_letter(total_col)
            avg_formula = f"=IFERROR({total_col_letter}{current_row}/COUNTIF({first_col_letter}{income_totals_row_num}:{last_col_letter}{income_totals_row_num},\"<>0\"),\"\")"
            ws.cell(row=current_row, column=avg_col, value=avg_formula)

            # Изменение = Excel формула тренда (SLOPE)
            trend_formula = make_trend_formula(first_col_letter, last_col_letter, current_row, len(months_list))
            change_cell = ws.cell(row=current_row, column=change_col, value=trend_formula)
            change_cell.number_format = '0.0%'

            # Применяем границы и чередование
            is_even = (current_row % 2 == 0)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border
                if is_even:
                    cell.fill = gray_fill

            income_category_rows[category] = current_row
            current_row += 1

        # Пустая строка перед ИТОГО ДОХОДЫ
        income_last_data_row = current_row - 1  # Последняя строка с данными категорий доходов
        current_row += 1  # Пропускаем строку

        # Строка ИТОГО ДОХОДЫ
        income_totals_row = current_row

        ws.cell(row=current_row, column=1, value='ИТОГО ДОХОДЫ' if lang == 'ru' else 'TOTAL INCOME')
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=None)

        # Для каждого месяца - формула SUM по колонке (только строки доходов)
        for col_offset, (year, month) in enumerate(months_list):
            col_idx = 3 + col_offset
            col_letter = get_column_letter(col_idx)
            is_current_month = (year == current_year and month == current_month)

            # Формула суммы по колонке (только строки доходов)
            sum_col_formula = f"=SUM({col_letter}{income_first_data_row}:{col_letter}{income_last_data_row})"
            cell = ws.cell(row=current_row, column=col_idx, value=sum_col_formula)
            if is_current_month:
                cell.font = Font(bold=True, color="0563C1")
            else:
                cell.font = Font(bold=True, color="008000")

        # Итого = SUM по строке
        first_col_letter = get_column_letter(first_month_col)
        last_col_letter = get_column_letter(last_month_col)
        sum_formula = f"=SUM({first_col_letter}{current_row}:{last_col_letter}{current_row})"
        ws.cell(row=current_row, column=total_col, value=sum_formula)
        ws.cell(row=current_row, column=total_col).font = Font(bold=True, color="008000")

        # Среднее для ИТОГО ДОХОДЫ = Итого / COUNTIF по своей строке (считаем непустые месяцы)
        total_col_letter = get_column_letter(total_col)
        avg_formula = f"=IFERROR({total_col_letter}{current_row}/COUNTIF({first_col_letter}{current_row}:{last_col_letter}{current_row},\"<>0\"),\"\")"
        ws.cell(row=current_row, column=avg_col, value=avg_formula)
        ws.cell(row=current_row, column=avg_col).font = Font(bold=True, color="008000")

        # Изменение % для ИТОГО ДОХОДЫ (Excel формула тренда)
        trend_formula = make_trend_formula(first_col_letter, last_col_letter, current_row, len(months_list))
        change_cell = ws.cell(row=current_row, column=change_col, value=trend_formula)
        change_cell.font = Font(bold=True, color="008000")
        change_cell.number_format = '0.0%'

        for col in range(1, len(headers) + 1):
            ws.cell(row=current_row, column=col).border = thin_border

        current_row += 2  # Пустая строка перед балансом

        # ==================== СТРОКА БАЛАНСА ====================
        balance_row = current_row
        ws.cell(row=current_row, column=1, value='БАЛАНС' if lang == 'ru' else 'BALANCE')
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=None)

        # Баланс = Доходы - Расходы (формула для каждого месяца)
        balance_values = []
        for col_offset, (year, month) in enumerate(months_list):
            col_idx = 3 + col_offset
            col_letter = get_column_letter(col_idx)
            is_current_month = (year == current_year and month == current_month)

            # Формула: ячейка ИТОГО ДОХОДЫ - ячейка ИТОГО РАСХОДЫ
            balance_formula = f"={col_letter}{income_totals_row}-{col_letter}{expenses_totals_row}"
            cell = ws.cell(row=current_row, column=col_idx, value=balance_formula)
            if is_current_month:
                cell.font = Font(bold=True, color="0563C1")
            else:
                cell.font = Font(bold=True)
            balance_values.append(income_month_totals[col_offset] - expense_month_totals[col_offset])

        # Итого = SUM по строке
        first_col_letter = get_column_letter(first_month_col)
        last_col_letter = get_column_letter(last_month_col)
        sum_formula = f"=SUM({first_col_letter}{current_row}:{last_col_letter}{current_row})"
        ws.cell(row=current_row, column=total_col, value=sum_formula)
        ws.cell(row=current_row, column=total_col).font = Font(bold=True)

        # Среднее для БАЛАНСА = Итого / COUNTIF по строке ИТОГО РАСХОДЫ (месяцы с данными)
        total_col_letter = get_column_letter(total_col)
        avg_formula = f"=IFERROR({total_col_letter}{current_row}/COUNTIF({first_col_letter}{expenses_totals_row}:{last_col_letter}{expenses_totals_row},\"<>0\"),\"\")"
        ws.cell(row=current_row, column=avg_col, value=avg_formula)
        ws.cell(row=current_row, column=avg_col).font = Font(bold=True)

        # Изменение для БАЛАНСА - оставляем пустым (не информативно)
        ws.cell(row=current_row, column=change_col, value=None).font = Font(bold=True)

        for col in range(1, len(headers) + 1):
            ws.cell(row=current_row, column=col).border = thin_border

        # ==================== НОРМА СБЕРЕЖЕНИЙ ====================
        savings_row = current_row + 1
        ws.cell(row=savings_row, column=1, value='Норма сбережений' if lang == 'ru' else 'Savings rate').font = Font(bold=True)
        ws.cell(row=savings_row, column=2, value=None)

        for col_offset, (year, month) in enumerate(months_list):
            col_idx = 3 + col_offset
            col_letter = get_column_letter(col_idx)
            formula = f"=IFERROR(({col_letter}{income_totals_row}-{col_letter}{expenses_totals_row})/{col_letter}{income_totals_row},\"\")"
            cell = ws.cell(row=savings_row, column=col_idx, value=formula)
            cell.number_format = '0.0%'
            cell.border = thin_border

        # Итого (по суммам)
        total_col_letter = get_column_letter(total_col)
        total_formula = f"=IFERROR(({total_col_letter}{income_totals_row}-{total_col_letter}{expenses_totals_row})/{total_col_letter}{income_totals_row},\"\")"
        total_cell = ws.cell(row=savings_row, column=total_col, value=total_formula)
        total_cell.number_format = '0.0%'
        total_cell.border = thin_border

        # Среднее – используем то же отношение суммарных доходов/расходов
        avg_col_letter = get_column_letter(avg_col)
        avg_formula = total_formula
        avg_cell = ws.cell(row=savings_row, column=avg_col, value=avg_formula)
        avg_cell.number_format = '0.0%'
        avg_cell.border = thin_border

        # Столбец изменения оставляем пустым
        ws.cell(row=savings_row, column=change_col, value=None).border = thin_border
        ws.cell(row=savings_row, column=1).border = thin_border
        ws.cell(row=savings_row, column=2).border = thin_border

        # Подсветка отклонений от плана (>+50% вверх и <50% вниз, но не 0)
        high_fill = PatternFill(start_color="FFF8E0E6", end_color="FFF8E0E6", fill_type="solid")  # мягкий розовый
        low_fill = PatternFill(start_color="FFE7F6EC", end_color="FFE7F6EC", fill_type="solid")   # мягкий зелёный

        def apply_plan_cf(start_row: int, end_row: int, invert: bool = False):
            if start_row > end_row:
                return
            for col_idx in range(first_month_col, first_month_col + len(months_list)):
                col_letter = get_column_letter(col_idx)
                rng = f"{col_letter}{start_row}:{col_letter}{end_row}"
                # Колонка B — план (фиксируем колонку, строка относительная), текущая колонка — факт
                formula_high = f"AND($B{start_row}>0,{col_letter}{start_row}>0,({col_letter}{start_row}-$B{start_row})/$B{start_row}>0.5)"
                formula_low = f"AND($B{start_row}>0,{col_letter}{start_row}>0,{col_letter}{start_row}/$B{start_row}<0.5)"
                high = low_fill if invert else high_fill
                low = high_fill if invert else low_fill
                ws.conditional_formatting.add(rng, FormulaRule(formula=[formula_high], fill=high))
                ws.conditional_formatting.add(rng, FormulaRule(formula=[formula_low], fill=low))

        # Расходы (перерасход = розовый, недобор к плану = зелёный)
        apply_plan_cf(3, last_data_row, invert=False)
        # Доходы (перевыполнение плана = зелёный, недовыполнение = розовый)
        apply_plan_cf(income_first_data_row, income_last_data_row, invert=True)

        # ==================== ДОП. АНАЛИТИКА: Топ-10 операций ====================
        def _prepare_top_ops(ops: Dict[Tuple[str, str], Dict[str, Any]]):
            items = []
            for op in ops.values():
                total_amount = op['total']
                total_count = op['count']
                avg_ticket = round(total_amount / total_count, 2) if total_count else 0
                inflation_ratio = calc_trend_python(op['monthly_totals'])
                # Округляем тренд до 4 знаков (это проценты: 0.1234 = 12.34%)
                if inflation_ratio is not None:
                    inflation_ratio = round(inflation_ratio, 4)
                items.append({
                    "description": op['description'],
                    "category": op['category'],
                    "currency": op['currency'],
                    "total": total_amount,
                    "count": total_count,
                    "avg": avg_ticket,
                    "inflation": inflation_ratio,
                })
            top_amount = sorted(items, key=lambda x: x["total"], reverse=True)[:10]
            top_count = sorted(items, key=lambda x: x["count"], reverse=True)[:10]
            return top_amount, top_count

        expense_top_amount, expense_top_count = _prepare_top_ops(expense_operations)
        income_top_amount, income_top_count = _prepare_top_ops(income_operations)

        def _render_top_table(start_row: int, title: str, items: list[dict]) -> int:
            ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)
            start_row += 1
            headers = [
                'Описание' if lang == 'ru' else 'Description',
                'Категория' if lang == 'ru' else 'Category',
                'Валюта' if lang == 'ru' else 'Currency',
                'Сумма' if lang == 'ru' else 'Amount',
                'Операций' if lang == 'ru' else 'Count',
                'Средний чек' if lang == 'ru' else 'Avg ticket',
                'Инфляция за год' if lang == 'ru' else 'Inflation YoY',
            ]
            for idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=start_row, column=idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            start_row += 1

            for item in items:
                ws.cell(row=start_row, column=1, value=item["description"])
                ws.cell(row=start_row, column=2, value=item["category"])
                ws.cell(row=start_row, column=3, value=item["currency"])
                amt_cell = ws.cell(row=start_row, column=4, value=smart_number(item["total"]))
                cnt_cell = ws.cell(row=start_row, column=5, value=item["count"] if item["count"] else None)
                avg_cell = ws.cell(row=start_row, column=6, value=smart_number(item["avg"]) if item["avg"] else None)
                infl_val = item["inflation"]
                infl_cell = ws.cell(row=start_row, column=7, value=infl_val if infl_val is not None else None)
                amt_cell.number_format = 'General'
                if avg_cell.value is not None:
                    avg_cell.number_format = 'General'
                if infl_val is not None:
                    infl_cell.number_format = '0.0%'
                for col_idx in range(1, 8):
                    ws.cell(row=start_row, column=col_idx).border = thin_border
                start_row += 1
            return start_row

        def _render_top_table(start_row: int, start_col: int, title: str, items: list[dict], inflation_header: str, header_fill_param: PatternFill) -> int:
            ws.cell(row=start_row, column=start_col, value=title).font = Font(bold=True, size=12)
            start_row += 1
            headers = [
                'Описание' if lang == 'ru' else 'Description',
                'Категория' if lang == 'ru' else 'Category',
                'Валюта' if lang == 'ru' else 'Currency',
                'Сумма' if lang == 'ru' else 'Amount',
                'Операций' if lang == 'ru' else 'Count',
                'Средний чек' if lang == 'ru' else 'Avg ticket',
                inflation_header,
            ]
            for idx, header in enumerate(headers, start=0):
                cell = ws.cell(row=start_row, column=start_col + idx, value=header)
                cell.font = header_font
                cell.fill = header_fill_param
                cell.alignment = header_alignment
                cell.border = thin_border
            start_row += 1

            for item in items:
                ws.cell(row=start_row, column=start_col + 0, value=item["description"])
                ws.cell(row=start_row, column=start_col + 1, value=item["category"])
                ws.cell(row=start_row, column=start_col + 2, value=item["currency"])
                amt_cell = ws.cell(row=start_row, column=start_col + 3, value=smart_number(item["total"]))
                cnt_cell = ws.cell(row=start_row, column=start_col + 4, value=item["count"] if item["count"] else None)
                avg_cell = ws.cell(row=start_row, column=start_col + 5, value=smart_number(item["avg"]) if item["avg"] else None)
                infl_val = item["inflation"]
                infl_cell = ws.cell(row=start_row, column=start_col + 6, value=infl_val if infl_val is not None else None)
                amt_cell.number_format = 'General'
                if avg_cell.value is not None:
                    avg_cell.number_format = 'General'
                if infl_val is not None:
                    infl_cell.number_format = '0.0%'
                for col_idx in range(start_col, start_col + 7):
                    ws.cell(row=start_row, column=col_idx).border = thin_border
                start_row += 1
            return start_row

        analytics_start_row = 1  # на одном уровне с основной таблицей, справа
        tables_start_col = change_col + 2  # размещаем справа от основной таблицы
        exp_infl_header = 'Инфл/год' if lang == 'ru' else 'Infl./yr'
        inc_infl_header = 'Измен/год' if lang == 'ru' else 'Change/yr'

        next_row = _render_top_table(
            analytics_start_row,
            tables_start_col,
            'Топ-10 расходов по сумме' if lang == 'ru' else 'Top-10 expenses by amount',
            expense_top_amount,
            exp_infl_header,
            header_fill
        )
        next_row = _render_top_table(
            next_row + 1,
            tables_start_col,
            'Топ-10 расходов по частоте' if lang == 'ru' else 'Top-10 expenses by count',
            expense_top_count,
            exp_infl_header,
            header_fill
        )

        next_row = _render_top_table(
            next_row + 2,
            tables_start_col,
            'Топ-10 доходов по сумме' if lang == 'ru' else 'Top-10 incomes by amount',
            income_top_amount,
            inc_infl_header,
            income_header_fill
        )
        _render_top_table(
            next_row + 1,
            tables_start_col,
            'Топ-10 доходов по частоте' if lang == 'ru' else 'Top-10 incomes by count',
            income_top_count,
            inc_infl_header,
            income_header_fill
        )

        # ==================== АВТОПОДБОР ШИРИНЫ КОЛОНОК ====================
        def autosize_columns(worksheet):
            for col_idx in range(1, worksheet.max_column + 1):
                max_len = 0
                for row_idx in range(1, worksheet.max_row + 1):
                    val = worksheet.cell(row=row_idx, column=col_idx).value
                    if val is not None:
                        max_len = max(max_len, len(str(val)))
                if max_len > 0:
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

        autosize_columns(ws)

        # Поджимаем узкие столбцы для процентов
        narrow_headers = {'Инфл/год', 'Infl./yr', 'Измен/год', 'Change/yr'}
        for c in range(1, ws.max_column + 1):
            header_val = ws.cell(row=analytics_start_row + 1, column=c).value  # строка с заголовками топов справа
            if header_val in narrow_headers or c == change_col:
                ws.column_dimensions[get_column_letter(c)].width = 10

        # Ограничиваем ширины основных столбцов таблицы месяцев
        ws.column_dimensions[get_column_letter(1)].width = min(ws.column_dimensions[get_column_letter(1)].width or 0, 30)  # Категория
        ws.column_dimensions[get_column_letter(2)].width = min(ws.column_dimensions[get_column_letter(2)].width or 0, 10)  # План
        for col_idx in range(3, 3 + len(months_list)):
            ws.column_dimensions[get_column_letter(col_idx)].width = min(ws.column_dimensions[get_column_letter(col_idx)].width or 0, 9)  # Месяцы
        ws.column_dimensions[get_column_letter(total_col)].width = min(ws.column_dimensions[get_column_letter(total_col)].width or 0, 12)
        ws.column_dimensions[get_column_letter(avg_col)].width = min(ws.column_dimensions[get_column_letter(avg_col)].width or 0, 10)
        ws.column_dimensions[get_column_letter(change_col)].width = min(ws.column_dimensions[get_column_letter(change_col)].width or 0, 10)

        # Возвращаем информацию для добавления гиперссылок позже
        return {
            'expense_rows': expense_category_rows,
            'income_rows': income_category_rows,
            'expenses_totals_row': expenses_totals_row,
            'income_totals_row': income_totals_row,
            'balance_row': balance_row,
            'current_month_col': 2 + len(months_list),  # Колонка текущего месяца
            'current_sheet_name': current_sheet_name
        }

    @staticmethod
    def generate_xlsx_with_charts(
        expenses: List[Expense],
        incomes: List[Income],
        year: int,
        month: int,
        user_id: int,
        lang: str = 'ru',
        household_mode: bool = False
    ) -> BytesIO:
        """
        Генерация XLSX файла с операциями, сводкой и графиками на одном листе.
        Структура: Траты слева (A-G), Summary справа (I-N), Графики правее (P+)

        Args:
            expenses: Список трат
            incomes: Список доходов
            year: Год
            month: Месяц
            user_id: ID пользователя
            lang: Язык (ru/en)
            household_mode: Режим семейного бюджета

        Returns:
            BytesIO объект с XLSX файлом
        """
        operations = ExportService.prepare_operations_data(expenses, incomes, lang)
        category_cashbacks = ExportService.calculate_category_cashbacks(expenses, user_id, month, household_mode)

        # Рассчитываем кешбэк для каждой траты
        expense_cashbacks = {}  # {expense_id: cashback_amount}
        has_any_cashback = False

        if expenses:
            # Рассчитываем кешбэк для каждой траты
            for expense in expenses:
                if expense.category_id and expense.category_id in category_cashbacks:
                    category_total = sum(float(e.amount) for e in expenses if e.category_id == expense.category_id)
                    if category_total > 0:
                        cashback_amount = float(expense.amount) * (category_cashbacks[expense.category_id] / category_total)
                        expense_cashbacks[expense.id] = cashback_amount
                        if cashback_amount > 0:
                            has_any_cashback = True

        wb = Workbook()
        ws = wb.active

        # Функция для обрезания длинного текста с многоточием
        def truncate_text(text: str, max_length: int = 33) -> str:
            """Обрезает текст до max_length символов, добавляя многоточие если нужно
            max_length=33 соответствует длине "Коммунальные услуги и подписки"
            """
            if len(text) > max_length:
                return text[:max_length-1] + '…'
            return text

        # ==================== НАЗВАНИЕ ЛИСТА В ФОРМАТЕ Oct-2025 ====================
        month_names_en = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                         'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        month_names_ru = ['Янв', 'Фев', 'Мар', 'Апр', 'Май', 'Июн',
                         'Июл', 'Авг', 'Сен', 'Окт', 'Ноя', 'Дек']

        if lang == 'ru':
            ws.title = f"{month_names_ru[month-1]}-{year}"
        else:
            ws.title = f"{month_names_en[month-1]}-{year}"

        # Стили
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")  # Мягкий синий
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Стили для зебра-полосок и границ
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # ==================== ЛЕВАЯ ЧАСТЬ: ТРАТЫ (Колонки A-G или A-H если есть кешбэк) ====================
        # Заголовок секции
        operations_section_title = 'ДНЕВНИК ОПЕРАЦИЙ' if lang == 'ru' else 'OPERATIONS LOG'
        ops_title_cell = ws.cell(row=1, column=1, value=operations_section_title)
        ops_title_cell.font = Font(bold=True, color="FFFFFF", size=12)
        ops_title_cell.fill = PatternFill(start_color="4A7BB7", end_color="4A7BB7", fill_type="solid")  # Темно-синий мягкий
        ops_title_cell.alignment = header_alignment
        ops_title_cell.border = thin_border

        # Порядок: Дата, Время, Сумма, Кешбэк (если есть), Валюта, Категория, Описание, Тип
        if has_any_cashback:
            if lang == 'en':
                headers_left = ['Date', 'Time', 'Amount', 'Cashback', 'Currency', 'Category', 'Description', 'Type']
            else:
                headers_left = ['Дата', 'Время', 'Сумма', 'Кешбэк', 'Валюта', 'Категория', 'Описание', 'Тип']
        else:
            if lang == 'en':
                headers_left = ['Date', 'Time', 'Amount', 'Currency', 'Category', 'Description', 'Type']
            else:
                headers_left = ['Дата', 'Время', 'Сумма', 'Валюта', 'Категория', 'Описание', 'Тип']

        ops_end_col = len(headers_left)  # 8 если есть кешбэк, 7 если нет

        for idx, header in enumerate(headers_left, start=1):
            cell = ws.cell(row=2, column=idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Заполнение данных трат
        current_row = 3  # Начинаем с 3-й строки (1 - заголовок секции, 2 - заголовки колонок)
        expenses_by_currency = {}
        incomes_by_currency = {}
        total_cashback_by_currency = {}  # Для подсчета общего кешбэка по валютам

        for op in operations:
            type_text = get_text('income', lang) if op['type'] == 'income' else get_text('expense', lang)

            # Санитизация
            description = str(op['description'] or '').replace('\n', ' ').replace('\r', ' ').strip()
            category = str(op['category'] or '').replace('\n', ' ').replace('\r', ' ').strip()

            # Получаем кешбэк для траты
            cashback = 0
            if op['type'] == 'expense' and 'object' in op and hasattr(op['object'], 'id'):
                cashback = expense_cashbacks.get(op['object'].id, 0)

            # Заполняем колонки A-G или A-H (в зависимости от has_any_cashback)
            ws.cell(row=current_row, column=1, value=op['date'].strftime('%d.%m.%Y'))
            ws.cell(row=current_row, column=2, value=op['time'].strftime('%H:%M'))
            ws.cell(row=current_row, column=3, value=smart_number(op['amount']))

            if has_any_cashback:
                # С колонкой кешбэка: Дата, Время, Сумма, Кешбэк, Валюта, Категория, Описание, Тип
                ws.cell(row=current_row, column=4, value=smart_number(cashback) if op['type'] == 'expense' else None)
                ws.cell(row=current_row, column=5, value=op['currency'])
                ws.cell(row=current_row, column=6, value=category)
                ws.cell(row=current_row, column=7, value=description)
                ws.cell(row=current_row, column=8, value=type_text)
            else:
                # Без колонки кешбэка: Дата, Время, Сумма, Валюта, Категория, Описание, Тип
                ws.cell(row=current_row, column=4, value=op['currency'])
                ws.cell(row=current_row, column=5, value=category)
                ws.cell(row=current_row, column=6, value=description)
                ws.cell(row=current_row, column=7, value=type_text)

            # Применяем границы и чередующуюся заливку к строке
            is_even_row = (current_row % 2 == 0)
            for col in range(1, ops_end_col + 1):  # Колонки A-G или A-H
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border
                if is_even_row:
                    cell.fill = gray_fill

            # Форматирование суммы
            amount_cell = ws.cell(row=current_row, column=3)
            if op['type'] == 'income':
                amount_cell.font = Font(color="008000", bold=True)
            else:
                amount_cell.font = Font(color="000000")  # Черный цвет для трат
            amount_cell.number_format = 'General'

            # Форматирование кешбэка (зеленым цветом)
            if has_any_cashback and cashback > 0:
                cashback_cell = ws.cell(row=current_row, column=4)
                cashback_cell.font = Font(color="008000", bold=True)
                cashback_cell.number_format = 'General'

            # Подсчет по валютам
            currency = op['currency']
            if op['type'] == 'expense':
                if currency not in expenses_by_currency:
                    expenses_by_currency[currency] = 0
                expenses_by_currency[currency] += abs(op['amount'])

                # Подсчет кешбэка по валютам
                if cashback > 0:
                    if currency not in total_cashback_by_currency:
                        total_cashback_by_currency[currency] = 0
                    total_cashback_by_currency[currency] += cashback
            else:
                if currency not in incomes_by_currency:
                    incomes_by_currency[currency] = 0
                incomes_by_currency[currency] += op['amount']

            current_row += 1

        expenses_data_end_row = current_row - 1

        # Добавляем автофильтр для таблицы операций (интерактивная сортировка и фильтрация)
        if expenses_data_end_row > 2:  # Если есть хотя бы одна операция
            ws.auto_filter.ref = f"A2:{get_column_letter(ops_end_col)}{expenses_data_end_row}"

        # Итоги (Расходы, Доходы, Кешбэк, Баланс)
        current_row += 1
        all_currencies = set(list(expenses_by_currency.keys()) + list(incomes_by_currency.keys()))

        for currency in sorted(all_currencies):
            expense_total = expenses_by_currency.get(currency, 0)
            income_total = incomes_by_currency.get(currency, 0)
            cashback_total = total_cashback_by_currency.get(currency, 0)
            balance = income_total - expense_total

            # Расходы (ВСЕГДА показываем, даже если 0)
            label = 'Расходы:' if lang == 'ru' else 'Expenses:'
            ws.cell(row=current_row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=None)
            ws.cell(row=current_row, column=3, value=smart_number(-expense_total))
            if has_any_cashback:
                ws.cell(row=current_row, column=4, value=None)
                ws.cell(row=current_row, column=5, value=currency)
                for col in range(6, ops_end_col + 1):
                    ws.cell(row=current_row, column=col, value=None)
            else:
                ws.cell(row=current_row, column=4, value=currency)
                for col in range(5, ops_end_col + 1):
                    ws.cell(row=current_row, column=col, value=None)
            # Применяем границы к строке итогов
            for col in range(1, ops_end_col + 1):
                ws.cell(row=current_row, column=col).border = thin_border
            ws.cell(row=current_row, column=3).font = Font(bold=True, color="000000")  # Черный цвет
            ws.cell(row=current_row, column=3).number_format = 'General'
            current_row += 1

            # Доходы (ВСЕГДА показываем, даже если 0)
            label = 'Доходы:' if lang == 'ru' else 'Income:'
            ws.cell(row=current_row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=None)
            ws.cell(row=current_row, column=3, value=smart_number(income_total))
            if has_any_cashback:
                ws.cell(row=current_row, column=4, value=None)
                ws.cell(row=current_row, column=5, value=currency)
                for col in range(6, ops_end_col + 1):
                    ws.cell(row=current_row, column=col, value=None)
            else:
                ws.cell(row=current_row, column=4, value=currency)
                for col in range(5, ops_end_col + 1):
                    ws.cell(row=current_row, column=col, value=None)
            # Применяем границы к строке итогов
            for col in range(1, ops_end_col + 1):
                ws.cell(row=current_row, column=col).border = thin_border
            ws.cell(row=current_row, column=3).font = Font(bold=True, color="008000")
            ws.cell(row=current_row, column=3).number_format = 'General'
            current_row += 1

            # Кешбэк (показываем только если есть)
            if cashback_total > 0:
                label = 'Кешбэк:' if lang == 'ru' else 'Cashback:'
                ws.cell(row=current_row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=current_row, column=2, value=None)
                ws.cell(row=current_row, column=3, value=smart_number(cashback_total))
                if has_any_cashback:
                    ws.cell(row=current_row, column=4, value=None)
                    ws.cell(row=current_row, column=5, value=currency)
                    for col in range(6, ops_end_col + 1):
                        ws.cell(row=current_row, column=col, value=None)
                else:
                    ws.cell(row=current_row, column=4, value=currency)
                    for col in range(5, ops_end_col + 1):
                        ws.cell(row=current_row, column=col, value=None)
                # Применяем границы к строке итогов
                for col in range(1, ops_end_col + 1):
                    ws.cell(row=current_row, column=col).border = thin_border
                ws.cell(row=current_row, column=3).font = Font(bold=True, color="008000")
                ws.cell(row=current_row, column=3).number_format = 'General'
                current_row += 1

            # Баланс (ВСЕГДА показываем)
            label = 'Баланс:' if lang == 'ru' else 'Balance:'
            ws.cell(row=current_row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=None)
            ws.cell(row=current_row, column=3, value=smart_number(balance))
            if has_any_cashback:
                ws.cell(row=current_row, column=4, value=None)
                ws.cell(row=current_row, column=5, value=currency)
                for col in range(6, ops_end_col + 1):
                    ws.cell(row=current_row, column=col, value=None)
            else:
                ws.cell(row=current_row, column=4, value=currency)
                for col in range(5, ops_end_col + 1):
                    ws.cell(row=current_row, column=col, value=None)
            # Применяем границы к строке итогов
            for col in range(1, ops_end_col + 1):
                ws.cell(row=current_row, column=col).border = thin_border
            ws.cell(row=current_row, column=3).font = Font(bold=True, color="0000FF")
            ws.cell(row=current_row, column=3).number_format = 'General'
            current_row += 2

        # Автоширина для колонок A-G или A-H
        for col in range(1, ops_end_col + 1):
            max_length = 0
            for row in range(1, current_row):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            # Столбец Cashback (4) делаем в 2 раза уже
            if col == 4 and has_any_cashback:
                ws.column_dimensions[get_column_letter(col)].width = min(max_length + 1, 50)
            else:
                ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 50)

        # ==================== ПРАВАЯ ЧАСТЬ: SUMMARY (Колонки K-P) ====================
        summary_start_col = 11  # Колонка K (отступ 2 столбца от дневника)

        # Подсчет статистики по категориям РАСХОДОВ
        category_stats = {}
        for op in operations:
            if op['type'] == 'expense':
                category_name = op['category'] or get_text('no_category', lang)
                currency = op['currency']
                amount = abs(op['amount'])

                key = (category_name, currency)
                if key not in category_stats:
                    category_stats[key] = {'total': 0, 'count': 0, 'category_ids': set()}

                category_stats[key]['total'] += amount
                category_stats[key]['count'] += 1

                # Сохраняем ВСЕ category_id (в household mode может быть несколько для одного названия)
                if 'object' in op and hasattr(op['object'], 'category_id'):
                    category_stats[key]['category_ids'].add(op['object'].category_id)

        # Инициализация переменных с дефолтными значениями (используются секцией доходов)
        charts_start_row = 3  # Если нет расходов, диаграммы доходов начинаются с 3-й строки
        dynamic_income_start_col = 30  # Колонка AD (минимальное значение по умолчанию)
        last_day = calendar.monthrange(year, month)[1]  # Количество дней в месяце
        sorted_days = list(range(0, last_day + 1))  # Список всех дней месяца (0-31)
        pie_block_height = 11.5  # Высота блока диаграммы (для расчета позиции таблицы)
        table_start_row = charts_start_row + int(pie_block_height + 15)  # Таблица под диаграммами

        # Подсчет расходов по дням и категориям (нужно ДО заполнения Summary для столбцов дней)
        daily_expenses_by_category = {}  # {category: {day: amount}}
        daily_cashback = {}  # Кешбек по дням

        # Получаем информацию о кешбеке для категорий (синхронно)
        category_cashback_rates = {}
        cashbacks_for_rates = Cashback.objects.filter(profile__telegram_id=user_id)
        for cb in cashbacks_for_rates:
            if cb.category_id:
                category_cashback_rates[cb.category_id] = float(cb.cashback_percent) / 100.0

        for op in operations:
            if op['type'] == 'expense':
                category = op['category'] or get_text('no_category', lang)
                day = op['date'].day
                amount = abs(op['amount'])

                # Расходы по категориям и дням
                if category not in daily_expenses_by_category:
                    daily_expenses_by_category[category] = {}
                if day not in daily_expenses_by_category[category]:
                    daily_expenses_by_category[category][day] = 0
                daily_expenses_by_category[category][day] += amount

                # Кешбек по дням
                cat_id = op.get('category_id')
                if cat_id and cat_id in category_cashback_rates:
                    cashback_rate = category_cashback_rates[cat_id]
                    day_cashback = amount * cashback_rate
                    if day not in daily_cashback:
                        daily_cashback[day] = 0
                    daily_cashback[day] += day_cashback

        # ЕСЛИ ЕСТЬ РАСХОДЫ - создаем секцию расходов (заголовки, summary, диаграммы)
        if category_stats:
            sorted_categories = sorted(
                category_stats.items(), key=lambda x: x[1]['total'], reverse=True
            )
            total_expenses = sum(stats['total'] for _, stats in sorted_categories)
            pie_segment_ratios: List[float] = []

            # Заголовок секции расходов
            expenses_section_title = 'РАСХОДЫ' if lang == 'ru' else 'EXPENSES'
            exp_title_cell = ws.cell(row=1, column=summary_start_col, value=expenses_section_title)
            exp_title_cell.font = Font(bold=True, color="FFFFFF", size=12)
            exp_title_cell.fill = PatternFill(start_color="C85A54", end_color="C85A54", fill_type="solid")  # Терракотовый красный мягкий
            exp_title_cell.alignment = header_alignment
            exp_title_cell.border = thin_border

            # Заголовки Summary
            if lang == 'en':
                headers_right = ['Category', 'Currency', 'Total', 'Count', 'Average', 'Cashback']
            else:
                headers_right = ['Категория', 'Валюта', 'Всего', 'Количество', 'Средний чек', 'Кешбэк']

            for idx, header in enumerate(headers_right):
                cell = ws.cell(row=2, column=summary_start_col + idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Добавляем заголовки дней (1-31) после Кешбэк - ВСЕ числа для таблицы
            days_start_col = summary_start_col + 6  # После Кешбэк
            for day_num in range(1, last_day + 1):
                cell = ws.cell(row=2, column=days_start_col + day_num - 1, value=day_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                # Ширина колонок дней устанавливается позже в блоке автоширины

            # Метки для диаграммы берутся из строки 2 (заголовки дней 1, 2, 3...)
            # Не записываем отдельные метки 5, 10, 15... в строку 1 - они дублируют заголовки
            chart_labels_row = 2  # Используем строку 2 с номерами дней для меток диаграммы

            # Заполнение Summary
            summary_row = 3  # Начинаем с 3-й строки (1 - заголовок секции, 2 - заголовки колонок)
            for (category, currency), stats in sorted_categories:
                average = stats['total'] / stats['count'] if stats['count'] > 0 else 0

                # Кешбэк - СУММА по ВСЕМ category_id для этой категории (важно для household mode!)
                total_cashback = 0
                for category_id in stats.get('category_ids', set()):
                    total_cashback += category_cashbacks.get(category_id, 0)
                cashback = total_cashback

                # Заполняем колонки Summary
                ws.cell(row=summary_row, column=summary_start_col, value=truncate_text(category))
                ws.cell(row=summary_row, column=summary_start_col + 1, value=currency)
                ws.cell(row=summary_row, column=summary_start_col + 2, value=smart_number(stats['total']))
                ws.cell(row=summary_row, column=summary_start_col + 3, value=stats['count'])
                ws.cell(row=summary_row, column=summary_start_col + 4, value=smart_number(average))
                ws.cell(row=summary_row, column=summary_start_col + 5, value=smart_number(cashback))

                # Заполняем данные по дням для этой категории
                for day_num in range(1, last_day + 1):
                    day_amount = daily_expenses_by_category.get(category, {}).get(day_num, None)
                    cell = ws.cell(row=summary_row, column=days_start_col + day_num - 1, value=smart_number(day_amount) if day_amount else None)
                    if day_amount:
                        cell.number_format = 'General'

                # Применяем границы и чередующуюся заливку к строке Summary (включая дни)
                is_even_row = (summary_row % 2 == 0)
                for col in range(summary_start_col, days_start_col + last_day):
                    cell = ws.cell(row=summary_row, column=col)
                    cell.border = thin_border
                    if is_even_row:
                        cell.fill = gray_fill

                # Форматирование
                ws.cell(row=summary_row, column=summary_start_col + 2).number_format = 'General'
                ws.cell(row=summary_row, column=summary_start_col + 4).number_format = 'General'
                ws.cell(row=summary_row, column=summary_start_col + 5).number_format = 'General'

                # Кешбэк зеленым если > 0
                if cashback > 0:
                    ws.cell(row=summary_row, column=summary_start_col + 5).font = Font(color="008000", bold=True)

                ratio = (stats['total'] / total_expenses) if total_expenses else 0
                pie_segment_ratios.append(ratio)

                summary_row += 1

            summary_end_row = summary_row - 1

            # Добавляем строку ИТОГО
            totals_row = summary_row
            ws.cell(row=totals_row, column=summary_start_col, value='ИТОГО' if lang == 'ru' else 'TOTAL')
            ws.cell(row=totals_row, column=summary_start_col).font = Font(bold=True)
            ws.cell(row=totals_row, column=summary_start_col + 1, value='')  # Валюта - пусто
            # Сумма всех расходов
            total_all = sum(stats['total'] for _, stats in sorted_categories)
            ws.cell(row=totals_row, column=summary_start_col + 2, value=smart_number(total_all))
            ws.cell(row=totals_row, column=summary_start_col + 2).number_format = 'General'
            ws.cell(row=totals_row, column=summary_start_col + 2).font = Font(bold=True)
            # Количество всех операций
            total_count = sum(stats['count'] for _, stats in sorted_categories)
            ws.cell(row=totals_row, column=summary_start_col + 3, value=total_count)
            ws.cell(row=totals_row, column=summary_start_col + 3).font = Font(bold=True)
            # Средний чек общий
            avg_all = total_all / total_count if total_count > 0 else 0
            ws.cell(row=totals_row, column=summary_start_col + 4, value=smart_number(avg_all))
            ws.cell(row=totals_row, column=summary_start_col + 4).number_format = 'General'
            ws.cell(row=totals_row, column=summary_start_col + 4).font = Font(bold=True)
            # Сумма всех кешбэков
            total_cashback_sum = sum(category_cashbacks.get(cat_id, 0) for (_, _), stats in sorted_categories for cat_id in stats.get('category_ids', set()))
            ws.cell(row=totals_row, column=summary_start_col + 5, value=smart_number(total_cashback_sum))
            ws.cell(row=totals_row, column=summary_start_col + 5).number_format = 'General'
            if total_cashback_sum > 0:
                ws.cell(row=totals_row, column=summary_start_col + 5).font = Font(color="008000", bold=True)
            else:
                ws.cell(row=totals_row, column=summary_start_col + 5).font = Font(bold=True)

            # Итоги по дням
            for day_num in range(1, last_day + 1):
                day_total = sum(daily_expenses_by_category.get(cat, {}).get(day_num, 0) for cat in daily_expenses_by_category)
                cell = ws.cell(row=totals_row, column=days_start_col + day_num - 1, value=smart_number(day_total) if day_total else None)
                if day_total:
                    cell.number_format = 'General'
                    cell.font = Font(bold=True)

            # Применяем границы к строке ИТОГО
            for col in range(summary_start_col, days_start_col + last_day):
                ws.cell(row=totals_row, column=col).border = thin_border

            summary_row += 1

            # Автоширина для всех колонок расходов (Summary + дни)
            for col in range(summary_start_col, days_start_col + last_day):
                max_length = 0
                for row in range(1, summary_row + 1):  # Включая ИТОГО
                    cell = ws.cell(row=row, column=col)
                    if cell.value:
                        # Для чисел форматируем с разделителями
                        val = cell.value
                        if isinstance(val, (int, float)):
                            formatted = f"{val:,.2f}" if isinstance(val, float) else str(val)
                            max_length = max(max_length, len(formatted))
                        else:
                            max_length = max(max_length, len(str(val)))

                col_offset = col - summary_start_col
                if col_offset == 0:  # Категория - шире
                    ws.column_dimensions[get_column_letter(col)].width = min(max(max_length + 3, 18), 45)
                elif col_offset == 1:  # Валюта
                    ws.column_dimensions[get_column_letter(col)].width = min(max(max_length + 2, 8), 12)
                elif col_offset == 2:  # Всего - числа с разделителями
                    ws.column_dimensions[get_column_letter(col)].width = min(max(max_length + 3, 12), 20)
                elif col_offset == 3:  # Количество
                    ws.column_dimensions[get_column_letter(col)].width = min(max(max_length + 2, 10), 14)
                elif col_offset == 4:  # Средний чек
                    ws.column_dimensions[get_column_letter(col)].width = min(max(max_length + 3, 12), 18)
                elif col_offset == 5:  # Кешбэк
                    ws.column_dimensions[get_column_letter(col)].width = min(max(max_length + 2, 10), 14)
                else:  # Дни - компактно но читаемо
                    ws.column_dimensions[get_column_letter(col)].width = min(max(max_length + 1, 6), 12)

            total_expenses = sum(stats['total'] for _, stats in sorted_categories)
            pie_segment_ratios = []
            for (_, _), stats in sorted_categories:
                amount = stats['total']
                ratio = (amount / total_expenses) if total_expenses else 0
                pie_segment_ratios.append(ratio)
            # ==================== ГРАФИКИ ====================
            # Диаграммы размещаются ПОД таблицей Summary (вертикально)
            # Пересчитываем charts_start_row - после строки ИТОГО с отступом
            charts_start_row = totals_row + 2  # Начало диаграмм под ИТОГО с отступом в 1 строку
            # pie_block_height уже инициализирован выше

            # КРУГОВАЯ ДИАГРАММА ПО КАТЕГОРИЯМ (под таблицей Summary)
            if summary_end_row > 1:
                from openpyxl.chart.legend import Legend

                pie = PieChart()
                pie.title = "Расходы по категориям" if lang == 'ru' else "Expenses by Category"
                pie.varyColors = True
                pie.width = 19.3  # Немного уже для создания зазора со столбчатой диаграммой
                pie.height = 11.5488  # Блок выше для размещения заголовка
                pie.layout = Layout(
                    manualLayout=ManualLayout(
                        xMode="edge",
                        yMode="factor",
                        wMode="factor",
                        hMode="factor",
                        x=0.115,  # Сдвинули на 5% правее (было 0.065)
                        y=0.05,   # Стандартное положение по вертикали
                        w=0.42,
                        h=0.8
                    )
                )
                # Смещаем заголовок ближе к левому верхнему краю
                pie.title.layout = Layout(
                    manualLayout=ManualLayout(
                        xMode="edge",
                        yMode="edge",
                        x=0.02,
                        y=0.01
                    )
                )

                # Легенда внутри области диаграммы справа
                pie.legend = Legend()
                pie.legend.position = 'r'
                pie.legend.overlay = False  # Легенда не накладывается на диаграмму
                # Настройка позиции легенды - начало с 65%, увеличенная высота для одного столбца
                pie.legend.layout = Layout(
                    manualLayout=ManualLayout(
                        xMode="edge",
                        yMode="edge",
                        x=0.65,  # Начало легенды на 65% ширины
                        y=0.02,  # Минимальный отступ сверху
                        w=0.32,  # Ширина 32% для длинных категорий
                        h=0.96  # Максимальная высота - принудительно вертикальное размещение
                    )
                )

                # Данные: колонка категорий и колонка всего
                labels = Reference(ws, min_col=summary_start_col, min_row=3, max_row=summary_end_row)
                data = Reference(ws, min_col=summary_start_col + 2, min_row=2, max_row=summary_end_row)
                pie.add_data(data, titles_from_data=True)
                pie.set_categories(labels)

                pie_block_height = pie.height

                # Применяем цвета из палитры к каждому сегменту круговой диаграммы
                from openpyxl.chart.series import DataPoint
                if pie.series:
                    series = pie.series[0]
                    series.tx = None  # убираем ссылку на заголовок серии, чтобы не выводился Total/Ряд

                    data_labels = DataLabelList()
                    data_labels.showPercent = False
                    data_labels.showLeaderLines = False
                    data_labels.showValue = False
                    data_labels.showCatName = False
                    data_labels.showLegendKey = False
                    data_labels.showSerName = False
                    data_labels.showBubbleSize = False
                    data_labels.dLblPos = "bestFit"  # Автоматическое позиционирование для избежания перекрытий
                    data_labels.numFmt = "0%"
                    point_labels: List[DataLabel] = []
                    if RichText and ParagraphProperties and CharacterProperties:
                        data_labels.txPr = RichText(p=[
                            Paragraph(
                                pPr=ParagraphProperties(
                                    defRPr=CharacterProperties(sz=900, b=True, solidFill="FFFFFF")
                                )
                            )
                        ])
                    series.dLbls = data_labels

                    # Показываем подписи только для сегментов >= 3%
                    from openpyxl.chart.label import DataLabelList as DLL
                    for idx, ratio in enumerate(pie_segment_ratios):
                        if ratio >= 0.04:
                            point_label = DataLabel(
                                idx=idx,
                                showPercent=True,
                                showVal=False,
                                showCatName=False,
                                showSerName=False,
                                showLegendKey=False,
                                showLeaderLines=False
                            )
                            # Устанавливаем позицию для каждой метки
                            point_label.dLblPos = "bestFit"
                            point_labels.append(point_label)
                    if point_labels:
                        data_labels.dLbl = point_labels

                    for idx in range(summary_end_row - 1):  # Количество категорий
                        color_idx = idx % len(ExportService.CATEGORY_COLORS)
                        color_hex = ExportService.CATEGORY_COLORS[color_idx].lstrip('#').upper()

                        # Создаем точку данных с цветом
                        pt = DataPoint(idx=idx)
                        pt.graphicalProperties = GraphicalProperties(solidFill=color_hex)
                        series.dPt.append(pt)

                # Размещение ПОД таблицей Summary с гибким позиционированием
                pie_anchor_column = 11  # Колонка K
                pie_anchor = AnchorMarker(
                    col=pie_anchor_column - 1,
                    colOff=0,
                    row=max(charts_start_row - 1, 0),
                    rowOff=0
                )
                pie_extent = XDRPositiveSize2D(
                    cx=int(round(pie.width * CM_TO_EMU)),
                    cy=int(round(pie.height * CM_TO_EMU))
                )
                pie.anchor = OneCellAnchor(_from=pie_anchor, ext=pie_extent)
                ws.add_chart(pie)

            # СТОЛБЧАТАЯ ДИАГРАММА ПО ДНЯМ - берёт данные из объединённой таблицы Summary
            # Данные уже заполнены в столбцах дней (days_start_col + 0..last_day-1)
            # Строки: 3 до summary_end_row (категории), totals_row (ИТОГО)

            num_categories = len(sorted_categories)
            bar_chart_row = charts_start_row  # Используем тот же ряд что и круговая диаграмма

            if num_categories > 0 and last_day > 0:
                from openpyxl.chart.axis import ChartLines
                from openpyxl.drawing.line import LineProperties

                # СТОЛБЧАТАЯ ДИАГРАММА - один столбик на каждый день
                bar = BarChart()
                bar.type = "col"  # Вертикальные столбики
                bar.grouping = "clustered"  # Обычная группировка (не stacked)
                bar.gapWidth = 150  # Расстояние между столбцами (больше = тоньше столбики)

                # Настраиваем заголовок и убираем подписи осей
                bar.title = "Расходы по дням" if lang == 'ru' else "Expenses by Day"
                bar.x_axis.title = None  # Убираем подпись оси X
                bar.y_axis.title = None  # Убираем подпись оси Y
                bar.legend = None  # Убираем легенду

                bar.width = 16  # Ширина диаграммы
                bar.height = 11.6  # Делаем немного выше для лучшего баланса

                # Настраиваем ось X (дни)
                # Метки 5, 10, 15, 20, 25, 30 задаются через пустые значения в заголовках
                bar.x_axis.tickLblSkip = 1  # Показываем все метки (пустые не видны)
                bar.x_axis.tickMarkSkip = 1
                bar.x_axis.delete = False  # Показываем ось X
                bar.x_axis.majorTickMark = "out"  # Метки снаружи

                # Настраиваем ось Y (суммы)
                bar.y_axis.delete = False  # Показываем ось Y с числами

                # Основные линии сетки Y (серые)
                bar.y_axis.majorGridlines = ChartLines()
                bar.y_axis.majorGridlines.spPr = GraphicalProperties(ln=LineProperties(solidFill="D0D0D0"))

                # Данные для диаграммы берём из объединённой таблицы:
                # Структура таблицы:
                # Строка 2: заголовки (Категория, Валюта, Всего, Кол-во, Средний, Кешбэк, 1, 2, 3... last_day)
                # Строки 3..summary_end_row: данные категорий
                #
                # Для stacked bar chart:
                # - Ось X: дни (1, 2, 3... last_day) - берём из строки 2, столбцы days_start_col..
                # - Серии: категории - каждая строка (3..summary_end_row) = отдельная серия
                # - Названия серий: из столбца summary_start_col (названия категорий)

                # Метки оси X - из строки 2 (заголовки дней 1, 2, 3...)
                days_labels = Reference(ws, min_col=days_start_col, max_col=days_start_col + last_day - 1, min_row=chart_labels_row)

                # Данные - ТОЛЬКО строка ИТОГО (totals_row)
                # Каждая ячейка в строке ИТОГО = значение для столбика этого дня
                # from_rows=True указывает что данные расположены в строке (горизонтально)
                data = Reference(ws,
                                 min_col=days_start_col,  # Начинаем с первого дня
                                 max_col=days_start_col + last_day - 1,  # До последнего дня
                                 min_row=totals_row,  # Строка ИТОГО
                                 max_row=totals_row)  # Только строка ИТОГО
                bar.add_data(data, from_rows=True, titles_from_data=False)
                bar.set_categories(days_labels)

                # Применяем цвета из палитры к каждой серии (категории)
                for idx, series in enumerate(bar.series):
                    color_hex = ExportService.CATEGORY_COLORS[idx % len(ExportService.CATEGORY_COLORS)].lstrip('#').upper()
                    series.graphicalProperties = GraphicalProperties(solidFill=color_hex)
                    series.dLbls = None

                # Настраиваем ось Y
                bar.y_axis.axId = 100
                bar.y_axis.crosses = "autoZero"

                # Размещение СПРАВА от круговой диаграммы (с отступом)
                bar_anchor_column = 20  # Колонка T - сдвинули правее для зазора между диаграммами
                bar_anchor = AnchorMarker(
                    col=bar_anchor_column - 1,
                    colOff=0,
                    row=max(bar_chart_row - 1, 0),
                    rowOff=0
                )
                bar_extent = XDRPositiveSize2D(
                    cx=int(round(bar.width * CM_TO_EMU)),
                    cy=int(round(bar.height * CM_TO_EMU))
                )
                bar.anchor = OneCellAnchor(_from=bar_anchor, ext=bar_extent)
                ws.add_chart(bar)

        # ==================== СЕКЦИЯ ДОХОДОВ (ПОД РАСХОДАМИ) ====================
        # Подсчет статистики по категориям доходов
        income_category_stats = {}
        for op in operations:
            if op['type'] == 'income':
                category_name = op['category'] or get_text('no_category', lang)
                currency = op['currency']
                amount = abs(op['amount'])

                key = (category_name, currency)
                if key not in income_category_stats:
                    income_category_stats[key] = {'total': 0, 'count': 0}

                income_category_stats[key]['total'] += amount
                income_category_stats[key]['count'] += 1

        # Подсчет доходов по дням и категориям (нужно ДО заполнения Summary)
        daily_incomes_by_category = {}  # {category: {day: amount}}
        for op in operations:
            if op['type'] == 'income':
                category = op['category'] or get_text('no_category', lang)
                day = op['date'].day
                amount = abs(op['amount'])

                if category not in daily_incomes_by_category:
                    daily_incomes_by_category[category] = {}
                if day not in daily_incomes_by_category[category]:
                    daily_incomes_by_category[category][day] = 0
                daily_incomes_by_category[category][day] += amount

        # Если есть доходы, создаем секцию доходов
        if income_category_stats:
            sorted_income_categories = sorted(
                income_category_stats.items(), key=lambda x: x[1]['total'], reverse=True
            )
            total_incomes = sum(stats['total'] for _, stats in sorted_income_categories)
            income_pie_segment_ratios: List[float] = []

            # Секция доходов размещается ПОД расходами (те же столбцы K+)
            # Вычисляем начальную строку для доходов ДИНАМИЧЕСКИ
            if category_stats:
                # Высота диаграммы в строках (примерно 1 строка = 0.5 см, диаграмма ~11.5 см = ~23 строки)
                chart_height_rows = 23
                # Начинаем после: таблицы расходов + ИТОГО + отступ + диаграммы + отступ
                # totals_row - последняя строка таблицы расходов
                # charts_start_row - где начинаются диаграммы
                income_section_start_row = charts_start_row + chart_height_rows + 2
            else:
                income_section_start_row = 1

            income_summary_start_col = summary_start_col  # Те же столбцы что и расходы (K)

            # Заголовки такие же как для расходов (включая Кешбэк для выравнивания столбцов)
            if lang == 'en':
                income_headers = ['Category', 'Currency', 'Total', 'Count', 'Average', 'Cashback']
            else:
                income_headers = ['Категория', 'Валюта', 'Всего', 'Количество', 'Средний', 'Кешбэк']

            # Заголовок секции доходов
            income_section_title = 'ДОХОДЫ' if lang == 'ru' else 'INCOME'
            title_cell = ws.cell(row=income_section_start_row, column=income_summary_start_col, value=income_section_title)
            title_cell.font = Font(bold=True, color="FFFFFF", size=12)
            title_cell.fill = PatternFill(start_color="5FA85F", end_color="5FA85F", fill_type="solid")  # Зеленый мягкий
            title_cell.alignment = header_alignment
            title_cell.border = thin_border

            # Заголовки колонок для Summary доходов
            income_header_row = income_section_start_row + 1
            for idx, header in enumerate(income_headers):
                cell = ws.cell(row=income_header_row, column=income_summary_start_col + idx, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color="7FB685", end_color="7FB685", fill_type="solid")  # Мятный зеленый
                cell.alignment = header_alignment
                cell.border = thin_border

            # Добавляем заголовки дней (1-31) после Кешбэк для доходов - ВСЕ числа для таблицы
            income_days_start_col = income_summary_start_col + 6  # После Кешбэк (6 колонок: Категория, Валюта, Всего, Кол-во, Средний, Кешбэк)
            for day_num in range(1, last_day + 1):
                cell = ws.cell(row=income_header_row, column=income_days_start_col + day_num - 1, value=day_num)
                cell.font = header_font
                cell.fill = PatternFill(start_color="7FB685", end_color="7FB685", fill_type="solid")
                cell.alignment = header_alignment
                cell.border = thin_border

            # Метки для диаграммы доходов берутся из строки заголовков (номера дней 1, 2, 3...)
            # Не записываем отдельные метки 5, 10, 15... - они дублируют заголовки
            income_chart_labels_row = income_header_row

            # Заполнение данных Summary для доходов
            income_summary_row = income_header_row + 1
            for (category, currency), stats in sorted_income_categories:
                average = stats['total'] / stats['count'] if stats['count'] > 0 else 0

                ws.cell(row=income_summary_row, column=income_summary_start_col, value=truncate_text(category))
                ws.cell(row=income_summary_row, column=income_summary_start_col + 1, value=currency)
                ws.cell(row=income_summary_row, column=income_summary_start_col + 2, value=smart_number(stats['total']))
                ws.cell(row=income_summary_row, column=income_summary_start_col + 3, value=stats['count'])
                ws.cell(row=income_summary_row, column=income_summary_start_col + 4, value=smart_number(average))
                ws.cell(row=income_summary_row, column=income_summary_start_col + 5, value='')  # Пустой Кешбэк для выравнивания с расходами

                # Заполняем данные по дням для этой категории доходов
                for day_num in range(1, last_day + 1):
                    day_amount = daily_incomes_by_category.get(category, {}).get(day_num, None)
                    cell = ws.cell(row=income_summary_row, column=income_days_start_col + day_num - 1, value=smart_number(day_amount) if day_amount else None)
                    if day_amount:
                        cell.number_format = 'General'

                # Применяем границы и чередующуюся заливку (включая дни)
                is_even_row = (income_summary_row % 2 == 0)
                for col in range(income_summary_start_col, income_days_start_col + last_day):
                    cell = ws.cell(row=income_summary_row, column=col)
                    cell.border = thin_border
                    if is_even_row:
                        cell.fill = gray_fill

                # Форматирование чисел
                ws.cell(row=income_summary_row, column=income_summary_start_col + 2).number_format = 'General'
                ws.cell(row=income_summary_row, column=income_summary_start_col + 4).number_format = 'General'

                ratio = (stats['total'] / total_incomes) if total_incomes else 0
                income_pie_segment_ratios.append(ratio)

                income_summary_row += 1

            income_summary_end_row = income_summary_row - 1

            # Добавляем строку ИТОГО для доходов
            income_totals_row = income_summary_row
            ws.cell(row=income_totals_row, column=income_summary_start_col, value='ИТОГО' if lang == 'ru' else 'TOTAL')
            ws.cell(row=income_totals_row, column=income_summary_start_col).font = Font(bold=True)
            ws.cell(row=income_totals_row, column=income_summary_start_col + 1, value='')  # Валюта - пусто
            # Сумма всех доходов
            income_total_all = sum(stats['total'] for _, stats in sorted_income_categories)
            ws.cell(row=income_totals_row, column=income_summary_start_col + 2, value=smart_number(income_total_all))
            ws.cell(row=income_totals_row, column=income_summary_start_col + 2).number_format = 'General'
            ws.cell(row=income_totals_row, column=income_summary_start_col + 2).font = Font(bold=True, color="008000")
            # Количество всех операций доходов
            income_total_count = sum(stats['count'] for _, stats in sorted_income_categories)
            ws.cell(row=income_totals_row, column=income_summary_start_col + 3, value=income_total_count)
            ws.cell(row=income_totals_row, column=income_summary_start_col + 3).font = Font(bold=True)
            # Средний доход общий
            income_avg_all = income_total_all / income_total_count if income_total_count > 0 else 0
            ws.cell(row=income_totals_row, column=income_summary_start_col + 4, value=smart_number(income_avg_all))
            ws.cell(row=income_totals_row, column=income_summary_start_col + 4).number_format = 'General'
            ws.cell(row=income_totals_row, column=income_summary_start_col + 4).font = Font(bold=True)
            ws.cell(row=income_totals_row, column=income_summary_start_col + 5, value='')  # Пустой Кешбэк для выравнивания

            # Итоги по дням для доходов
            for day_num in range(1, last_day + 1):
                day_total = sum(daily_incomes_by_category.get(cat, {}).get(day_num, 0) for cat in daily_incomes_by_category)
                cell = ws.cell(row=income_totals_row, column=income_days_start_col + day_num - 1, value=smart_number(day_total) if day_total else None)
                if day_total:
                    cell.number_format = 'General'
                    cell.font = Font(bold=True)

            # Применяем границы к строке ИТОГО доходов
            for col in range(income_summary_start_col, income_days_start_col + last_day):
                ws.cell(row=income_totals_row, column=col).border = thin_border

            # Диаграммы доходов размещаются ПОД таблицей доходов (как в расходах - сразу после ИТОГО)
            income_charts_start_row = income_totals_row + 2  # +2 как в расходах (summary_end_row + 2)

            # КРУГОВАЯ ДИАГРАММА ДОХОДОВ
            income_pie = PieChart()
            income_pie.title = "Доходы по категориям" if lang == 'ru' else "Income by Category"
            income_pie.varyColors = True
            income_pie.width = 19.3
            income_pie.height = 11.5488
            income_pie.layout = Layout(
                manualLayout=ManualLayout(
                    xMode="edge",
                    yMode="factor",
                    wMode="factor",
                    hMode="factor",
                    x=0.115,
                    y=0.05,
                    w=0.42,
                    h=0.8
                )
            )
            income_pie.title.layout = Layout(
                manualLayout=ManualLayout(
                    xMode="edge",
                    yMode="edge",
                    x=0.02,
                    y=0.01
                )
            )

            # Легенда внутри области диаграммы справа
            income_pie.legend = Legend()
            income_pie.legend.position = 'r'
            income_pie.legend.overlay = False
            income_pie.legend.layout = Layout(
                manualLayout=ManualLayout(
                    xMode="edge",
                    yMode="edge",
                    x=0.65,
                    y=0.02,
                    w=0.32,
                    h=0.96
                )
            )

            # Данные для круговой диаграммы доходов
            income_labels = Reference(ws, min_col=income_summary_start_col, min_row=income_header_row + 1, max_row=income_summary_end_row)
            income_data = Reference(ws, min_col=income_summary_start_col + 2, min_row=income_header_row, max_row=income_summary_end_row)
            income_pie.add_data(income_data, titles_from_data=True)
            income_pie.set_categories(income_labels)

            # Применяем цвета к сегментам круговой диаграммы доходов
            if income_pie.series:
                series = income_pie.series[0]
                series.tx = None

                data_labels = DataLabelList()
                data_labels.showPercent = False
                data_labels.showLeaderLines = False
                data_labels.showValue = False
                data_labels.showCatName = False
                data_labels.showLegendKey = False
                data_labels.showSerName = False
                data_labels.showBubbleSize = False
                data_labels.dLblPos = "bestFit"
                data_labels.numFmt = "0%"
                point_labels: List[DataLabel] = []
                if RichText and ParagraphProperties and CharacterProperties:
                    data_labels.txPr = RichText(p=[
                        Paragraph(
                            pPr=ParagraphProperties(
                                defRPr=CharacterProperties(sz=900, b=True, solidFill="FFFFFF")
                            )
                        )
                    ])
                series.dLbls = data_labels

                # Показываем подписи только для сегментов >= 4%
                for idx, ratio in enumerate(income_pie_segment_ratios):
                    if ratio >= 0.04:
                        point_label = DataLabel(
                            idx=idx,
                            showPercent=True,
                            showVal=False,
                            showCatName=False,
                            showSerName=False,
                            showLegendKey=False,
                            showLeaderLines=False
                        )
                        point_label.dLblPos = "bestFit"
                        point_labels.append(point_label)
                if point_labels:
                    data_labels.dLbl = point_labels

                for idx in range(income_summary_end_row - income_header_row):
                    color_idx = idx % len(ExportService.CATEGORY_COLORS)
                    color_hex = ExportService.CATEGORY_COLORS[color_idx].lstrip('#').upper()
                    pt = DataPoint(idx=idx)
                    pt.graphicalProperties = GraphicalProperties(solidFill=color_hex)
                    series.dPt.append(pt)

            # Размещение круговой диаграммы доходов под таблицей доходов
            income_pie_anchor = AnchorMarker(
                col=income_summary_start_col - 1,
                colOff=0,
                row=max(income_charts_start_row - 1, 0),
                rowOff=0
            )
            income_pie_extent = XDRPositiveSize2D(
                cx=int(round(income_pie.width * CM_TO_EMU)),
                cy=int(round(income_pie.height * CM_TO_EMU))
            )
            income_pie.anchor = OneCellAnchor(_from=income_pie_anchor, ext=income_pie_extent)
            ws.add_chart(income_pie)

            # СТОЛБЧАТАЯ ДИАГРАММА ДОХОДОВ ПО ДНЯМ
            num_income_categories = len(sorted_income_categories)
            if num_income_categories > 0 and last_day > 0:
                from openpyxl.chart.axis import ChartLines
                from openpyxl.drawing.line import LineProperties

                income_bar = BarChart()
                income_bar.type = "col"
                income_bar.grouping = "clustered"  # Обычная группировка (не stacked)
                income_bar.gapWidth = 150  # Расстояние между столбцами (больше = тоньше столбики)

                income_bar.title = "Доходы по дням" if lang == 'ru' else "Income by Day"
                income_bar.x_axis.title = None
                income_bar.y_axis.title = None
                income_bar.legend = None

                income_bar.width = 16
                income_bar.height = 11.6

                # Настраиваем ось X (дни)
                income_bar.x_axis.tickLblSkip = 1  # Показываем каждую метку
                income_bar.x_axis.tickMarkSkip = 1
                income_bar.x_axis.delete = False
                income_bar.x_axis.majorTickMark = "out"
                income_bar.y_axis.delete = False

                income_bar.y_axis.majorGridlines = ChartLines()
                income_bar.y_axis.majorGridlines.spPr = GraphicalProperties(ln=LineProperties(solidFill="D0D0D0"))

                # Данные для диаграммы доходов - ТОЛЬКО строка ИТОГО для простоты
                # Метки оси X - из скрытой строки (только 5, 10, 15, 20, 25, 30)
                income_days_labels = Reference(ws, min_col=income_days_start_col, max_col=income_days_start_col + last_day - 1, min_row=income_chart_labels_row)

                income_bar_data = Reference(ws,
                                          min_col=income_days_start_col,  # Начинаем с первого дня
                                          max_col=income_days_start_col + last_day - 1,  # До последнего дня
                                          min_row=income_totals_row,  # Строка ИТОГО
                                          max_row=income_totals_row)  # Только строка ИТОГО
                income_bar.add_data(income_bar_data, from_rows=True, titles_from_data=False)
                income_bar.set_categories(income_days_labels)

                # Применяем цвета
                for idx, series in enumerate(income_bar.series):
                    color_hex = ExportService.CATEGORY_COLORS[idx % len(ExportService.CATEGORY_COLORS)].lstrip('#').upper()
                    series.graphicalProperties = GraphicalProperties(solidFill=color_hex)
                    series.dLbls = None

                income_bar.y_axis.axId = 100
                income_bar.y_axis.crosses = "autoZero"

                # Размещение справа от круговой диаграммы доходов (с отступом)
                income_bar_anchor_column = 20  # Колонка T - как в расходах
                income_bar_anchor = AnchorMarker(
                    col=income_bar_anchor_column - 1,
                    colOff=0,
                    row=max(income_charts_start_row - 1, 0),
                    rowOff=0
                )
                income_bar_extent = XDRPositiveSize2D(
                    cx=int(round(income_bar.width * CM_TO_EMU)),
                    cy=int(round(income_bar.height * CM_TO_EMU))
                )
                income_bar.anchor = OneCellAnchor(_from=income_bar_anchor, ext=income_bar_extent)
                ws.add_chart(income_bar)

        # Закрепить заголовки (строки 1-2: заголовок секции + заголовки колонок)
        ws.freeze_panes = 'A3'

        # ==================== ДОБАВЛЯЕМ ЛИСТ "ИТОГИ 12 МЕС" ====================
        try:
            # Получаем profile для загрузки данных за 12 месяцев
            profile = Profile.objects.get(telegram_id=user_id)
            household_id = profile.household_id if household_mode else None

            # Добавляем лист с итогами (он станет первым листом)
            ExportService._add_summary_sheet(
                wb=wb,
                profile_id=profile.id,
                current_year=year,
                current_month=month,
                lang=lang,
                household_id=household_id
            )
        except Profile.DoesNotExist:
            logger.warning(f"Profile not found for user_id={user_id}, skipping summary sheet")
        except Exception as e:
            logger.error(f"Error adding summary sheet: {e}")

        # Сохранить в BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output
